from flask import Blueprint, abort, jsonify, render_template, request
from app import mysql
from io import BytesIO
from openpyxl import Workbook
from flask import send_file
from openpyxl.styles import Font, Alignment
import re
from collections import defaultdict
from flask_login import login_required


TABLAS_AFECCIONES = {
    "morbi": "afecciones_morbi",
    "morta": "afecciones_morta"
}

reportes = Blueprint('reportes', __name__, url_prefix='/reportes')


# =========================
# 1. VISTA (HTML)
# =========================
@reportes.route('/<tipo>/reporte')
@login_required
def reporte(tipo):

    tabla = TABLAS_AFECCIONES.get(tipo)

    if not tabla:
        abort(404, description="Tipo de reporte no válido.")

    titulo = "Reporte de Morbilidad" if tipo == "morbi" else "Reporte de Mortalidad"

    anios_disponibles = [2023, 2024, 2025, 2026]

    meses = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }

    cur = mysql.connection.cursor()

    # =========================
    # AFECCIONES
    # =========================
    cur.execute(f"""
        SELECT DISTINCT grupo
        FROM {tabla}
        ORDER BY grupo
    """)
    grupos = [r[0] for r in cur.fetchall()]

    # =========================
    # SERVICIOS DE EGRESO
    # =========================
    cur.execute(f"""
        SELECT DISTINCT especialidad
        FROM {tabla}
        WHERE especialidad IS NOT NULL
        AND especialidad <> ''
        ORDER BY especialidad
    """)

    servicio_disponibles = [r[0] for r in cur.fetchall()]

    # =========================
    # UNIDADES
    # =========================
    cur.execute(f"""
        SELECT DISTINCT c.nombre_unidad AS unidad
        FROM {tabla} m
        LEFT JOIN catalogo_unidades c
            ON m.clues = c.clues
        WHERE c.nombre_unidad IS NOT NULL
        ORDER BY unidad ASC
    """)

    unidades_disponibles = [
        r[0]
        for r in cur.fetchall()
    ]

    cur.close()

    return render_template(
        "reportes/reporte_afecciones.html",
        tipo=tipo,
        titulo=titulo,
        anios_disponibles=anios_disponibles,
        meses=meses,
        afecciones_disponibles=grupos,
        unidades_disponibles=unidades_disponibles,
        servicio_disponibles=servicio_disponibles
    )


# =========================
# 2. API JSON (REPORTE)
# =========================
@reportes.route('/<tipo>/data')
@login_required
def data_reporte(tipo):

    tabla = TABLAS_AFECCIONES.get(tipo)

    if not tabla:
        abort(404)


    def obtener_codigo(texto):
        m = re.match(r'^(\d+)', str(texto).strip())
        return int(m.group(1)) if m else 999999

    anios = request.args.get("anios")
    meses = request.args.get("meses")
    unidades = request.args.get("unidades")
    afecciones = request.args.get("afecciones")
    servicios = request.args.get("servicios")
    top = request.args.get("top", type=int)


    query = f"""
        SELECT
            a.anio,
            a.grupo AS causa,
            SUM(a.total) AS total
        FROM {tabla} a
        LEFT JOIN catalogo_unidades c
            ON a.clues = c.clues
        WHERE 1=1
    """

    params = []

    # ================= FILTROS =================
    if anios:
        lista_anios = anios.split(",")
        query += f" AND a.anio IN ({','.join(['%s'] * len(lista_anios))})"
        params.extend(lista_anios)

    if meses:
        lista_meses = meses.split(",")
        query += f" AND a.mes IN ({','.join(['%s'] * len(lista_meses))})"
        params.extend(lista_meses)

    if unidades:
        lista_unidades = unidades.split(",")
        query += f" AND c.nombre_unidad IN ({','.join(['%s'] * len(lista_unidades))})"
        params.extend(lista_unidades)

    if afecciones:
        lista_afecciones = afecciones.split(",")
        query += f" AND a.grupo IN ({','.join(['%s'] * len(lista_afecciones))})"
        params.extend(lista_afecciones)

    if servicios:
        lista_servicios = servicios.split(",")
        query += f" AND a.especialidad IN ({','.join(['%s'] * len(lista_servicios))})"
        params.extend(lista_servicios)

    query += """
        GROUP BY a.anio, a.grupo
        ORDER BY a.anio DESC, total DESC
    """

    cur = mysql.connection.cursor()
    cur.execute(query, params)
    rows = cur.fetchall()
    cur.close()

    # ==================================================
    # CASO ESPECIAL: UNA AFECCIÓN → DESGLOSE POR UNIDAD
    # ==================================================
    if afecciones and len(lista_afecciones) == 1:

        query_unidades = f"""
            SELECT
                a.anio,
                c.nombre_unidad,
                a.grupo,
                SUM(a.total) AS total
            FROM {tabla} a
            LEFT JOIN catalogo_unidades c
                ON a.clues = c.clues
            WHERE a.grupo = %s
        """

        params_unidades = [lista_afecciones[0]]

        if anios:
            query_unidades += f" AND a.anio IN ({','.join(['%s'] * len(lista_anios))})"
            params_unidades.extend(lista_anios)

        if meses:
            query_unidades += f" AND a.mes IN ({','.join(['%s'] * len(lista_meses))})"
            params_unidades.extend(lista_meses)

        if unidades:
            query_unidades += f" AND c.nombre_unidad IN ({','.join(['%s'] * len(lista_unidades))})"
            params_unidades.extend(lista_unidades)

        if servicios:
            query_unidades += f" AND a.especialidad IN ({','.join(['%s'] * len(lista_servicios))})"
            params_unidades.extend(lista_servicios)

        query_unidades += """
            GROUP BY a.anio, c.nombre_unidad, a.grupo
            ORDER BY a.anio DESC, total DESC
        """

        cur = mysql.connection.cursor()
        cur.execute(query_unidades, params_unidades)
        rows_unidades = cur.fetchall()
        cur.close()

        datos = defaultdict(list)
        totales = defaultdict(int)

        for r in rows_unidades:

            anio = r[0]

            datos[anio].append({
                "orden": None,
                "unidad": r[1] or "SIN UNIDAD",
                "causa": r[2],
                "total": int(r[3])
            })

            totales[anio] += int(r[3])

        resultado = []

        for anio, registros in datos.items():

            # ordenar por total descendente
            registros.sort(
                key=lambda x: x["total"],
                reverse=True
            )

            # aplicar TOP si existe
            if top:

                top_registros = registros[:top]

                resto = sum(
                    x["total"]
                    for x in registros[top:]
                )

                if resto > 0:
                    top_registros.append({
                        "orden": "",
                        "unidad": "",
                        "causa": "Resto de unidades",
                        "total": resto
                    })

                registros = top_registros

            # numerar
            contador = 1

            for item in registros:

                if item["causa"] != "Resto de unidades":
                    item["orden"] = contador
                    contador += 1
                else:
                    item["orden"] = ""

            resultado.append({
                "anio": anio,
                "total_unidad": totales[anio],
                "causas": registros
            })

        return jsonify(resultado)

    # ==================================================
    # CASO GENERAL
    # ==================================================
    data = defaultdict(list)
    totales_anio = defaultdict(int)

    for r in rows:

        anio, causa, total = r

        data[anio].append({
            "causa": causa,
            "total": total
        })

        totales_anio[anio] += total

    resultado = []

    for anio, causas in data.items():

        causas_normales = []
        mal_definidas = []
        las_demas = []

        for c in causas:

            nombre = str(c["causa"]).strip().upper()

            if nombre in ("LAS DEMAS", "LAS DEMÁS", "LAS DEMAS CAUSAS", "LAS DEMÁS CAUSAS"):
                las_demas.append(c)

            elif nombre in ("MAL DEFINIDAS", "MAL DEFINIDA"):
                mal_definidas.append(c)

            else:
                causas_normales.append(c)

        causas_normales.sort(
            key=lambda x: (-x["total"], obtener_codigo(x["causa"]))
        )

        # ================= TOP =================
        if top:
            top_causas = causas_normales[:top]
            resto = sum(x["total"] for x in causas_normales[top:])
        else:
            top_causas = causas_normales.copy()
            resto = 0

        resto += sum(x["total"] for x in mal_definidas)
        resto += sum(x["total"] for x in las_demas)

        contador = 1
        final_causas = []

        for item in top_causas:
            final_causas.append({
                "orden": contador,
                "causa": item["causa"],
                "total": int(item["total"])
            })
            contador += 1

        if resto > 0:
            final_causas.append({
                "orden": "",
                "causa": "Resto de causas",
                "total": int(resto)
            })

        resultado.append({
            "anio": anio,
            "total_unidad": totales_anio.get(anio, 0),
            "causas": final_causas or []
        })

    # ================= FALLBACK =================
    if not resultado:
        resultado = [{
            "anio": 0,
            "total_unidad": 0,
            "causas": []
        }]

    return jsonify(resultado)



@reportes.route('/<tipo>/excel')
@login_required
def excel_reporte(tipo):

    tabla = TABLAS_AFECCIONES.get(tipo)

    if not tabla:
        abort(404)

    datos = data_reporte(tipo).get_json()

    meses_nombre = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }

    meses_filtro = request.args.get("meses")
    unidades_filtro = request.args.get("unidades")
    servicios_filtro = request.args.get("servicios")

    wb = Workbook()
    ws = wb.active
    nombre = "Morbilidad" if tipo == "morbi" else "Mortalidad"

    ws.title = nombre

    fila = 1

    # =========================
    # TÍTULO
    # =========================
    ws.merge_cells(f"A{fila}:D{fila}")
    celda = ws.cell(fila, 1)

    celda.value = f"REPORTE DE {nombre.upper()}"
    celda.font = Font(bold=True, size=14)
    celda.alignment = Alignment(horizontal="center")

    fila += 2

    # =========================
    # UNIDADES
    # =========================
    if unidades_filtro:
        unidades = unidades_filtro.split(",")
        ws.cell(fila, 1, "Unidad(es):")
        ws.cell(fila, 2, ", ".join(unidades))
    else:
        ws.cell(fila, 1, "Unidad(es):")
        ws.cell(fila, 2, "Todas")

    fila += 2

    # =========================
    # SERVICIOS (NUEVO)
    # =========================
    ws.cell(fila, 1, "Servicio(s):")

    if servicios_filtro:
        servicios = servicios_filtro.split(",")
        ws.cell(fila, 2, ", ".join(servicios))
    else:
        ws.cell(fila, 2, "Todos")

    fila += 3

    # =========================
    # BLOQUES POR AÑO (HORIZONTAL)
    # =========================

    col_inicio = 1
    fila_base = fila

    for bloque in datos:

        fila = fila_base

        anio = bloque.get("anio", "N/A")

        lista_meses = []

        if meses_filtro:
            lista_meses = [
                int(x)
                for x in meses_filtro.split(",")
                if x.isdigit()
            ]

        periodo = str(anio)

        if lista_meses:
            periodo = (
                f"{meses_nombre.get(min(lista_meses), '')}"
                f" - "
                f"{meses_nombre.get(max(lista_meses), '')}"
                f" {anio}"
            )

        causas = bloque.get("causas", [])

        if not isinstance(causas, list):
            causas = []

        mostrar_unidad = (
            len(causas) > 0
            and isinstance(causas[0], dict)
            and "unidad" in causas[0]
        )

        # =========================
        # TITULO DEL AÑO
        # =========================

        if mostrar_unidad:
            ws.merge_cells(
                start_row=fila,
                start_column=col_inicio,
                end_row=fila,
                end_column=col_inicio + 3
            )
        else:
            ws.merge_cells(
                start_row=fila,
                start_column=col_inicio,
                end_row=fila,
                end_column=col_inicio + 2
            )

        celda = ws.cell(fila, col_inicio)
        celda.value = f"AÑO {anio}"
        celda.font = Font(bold=True, size=12)
        celda.alignment = Alignment(horizontal="center")

        fila += 1

        # =========================
        # PERIODO
        # =========================

        ws.cell(fila, col_inicio, "Periodo:")
        ws.cell(fila, col_inicio + 1, periodo)

        fila += 2

        # =========================
        # ENCABEZADOS
        # =========================

        ws.cell(
            fila,
            col_inicio,
            "Orden"
        ).font = Font(bold=True)

        if mostrar_unidad:

            ws.cell(
                fila,
                col_inicio + 1,
                "Unidad"
            ).font = Font(bold=True)

            ws.cell(
                fila,
                col_inicio + 2,
                "Causa"
            ).font = Font(bold=True)

            ws.cell(
                fila,
                col_inicio + 3,
                "Total"
            ).font = Font(bold=True)

        else:

            ws.cell(
                fila,
                col_inicio + 1,
                "Causa"
            ).font = Font(bold=True)

            ws.cell(
                fila,
                col_inicio + 2,
                "Total"
            ).font = Font(bold=True)

        fila += 1

        # =========================
        # DATOS
        # =========================

        for causa in causas:

            if not isinstance(causa, dict):
                continue

            if mostrar_unidad:

                ws.cell(
                    fila,
                    col_inicio,
                    causa.get("orden", "")
                )

                ws.cell(
                    fila,
                    col_inicio + 1,
                    causa.get("unidad", "")
                )

                ws.cell(
                    fila,
                    col_inicio + 2,
                    causa.get("causa", "SIN CAUSA")
                )

                ws.cell(
                    fila,
                    col_inicio + 3,
                    int(causa.get("total", 0) or 0)
                ).number_format = "#,##0"

            else:

                ws.cell(
                    fila,
                    col_inicio,
                    causa.get("orden", "")
                )

                ws.cell(
                    fila,
                    col_inicio + 1,
                    causa.get("causa", "SIN CAUSA")
                )

                ws.cell(
                    fila,
                    col_inicio + 2,
                    int(causa.get("total", 0) or 0)
                ).number_format = "#,##0"

            fila += 1

        # =========================
        # TOTAL GENERAL
        # =========================

        total_general = int(
            bloque.get("total_unidad", 0) or 0
        )

        if mostrar_unidad:

            ws.cell(
                fila,
                col_inicio + 2,
                "TOTAL GENERAL"
            ).font = Font(bold=True)

            ws.cell(
                fila,
                col_inicio + 3,
                total_general
            ).font = Font(bold=True)

            ws.cell(
                fila,
                col_inicio + 3
            ).number_format = "#,##0"

            # siguiente bloque
            col_inicio += 7

        else:

            ws.cell(
                fila,
                col_inicio + 1,
                "TOTAL GENERAL"
            ).font = Font(bold=True)

            ws.cell(
                fila,
                col_inicio + 2,
                total_general
            ).font = Font(bold=True)

            ws.cell(
                fila,
                col_inicio + 2
            ).number_format = "#,##0"

            # siguiente bloque
            col_inicio += 6

    # =========================
    # ANCHOS
    # =========================
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"Reporte_{nombre}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )